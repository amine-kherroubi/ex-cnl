from __future__ import annotations
from abc import ABC, abstractmethod
from typing import Any
import pandas as pd
from openpyxl import Workbook
from exceptions import FileProcessingError


class ExportStrategy(ABC):  # Strategy pattern
    @abstractmethod
    def export(self, data: dict[str, pd.DataFrame], output_path: str) -> None: ...


class ExcelExportStrategy(ExportStrategy):  # Strategy pattern implementation
    def export(self, data: dict[str, pd.DataFrame], output_path: str) -> None:
        try:
            workbook: Workbook = Workbook()
            workbook.remove(workbook.active)  # type: ignore

            for sheet_name, dataframe in data.items():
                self._create_sheet(workbook, sheet_name, dataframe)

            workbook.save(output_path)

        except Exception as error:
            raise FileProcessingError(f"Excel export failed: {error}") from error

    def _create_sheet(
        self, workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame
    ) -> None:
        worksheet: Any = workbook.create_sheet(title=sheet_name)

        # Add headers
        worksheet.append(dataframe.columns.tolist())

        # Add data rows
        for row_data in dataframe.itertuples(index=False, name=None):
            worksheet.append(list(row_data))


class ExportService(object):  # Service pattern with strategy
    def __init__(self, strategy: ExportStrategy) -> None:
        self._strategy: ExportStrategy = strategy

    def export_results(self, data: dict[str, pd.DataFrame], output_path: str) -> None:
        if not data:
            return

        self._strategy.export(data, output_path)
