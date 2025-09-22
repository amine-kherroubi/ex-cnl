from __future__ import annotations
from typing import Any, Literal, final

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


@final
class ExcelStyling(object):
    """Class to handle Excel worksheet styling with consistent formatting."""

    # Font class attributes
    FONT_NORMAL = Font(name="Arial", size=9, bold=False)
    FONT_BOLD = Font(name="Arial", size=9, bold=True)
    FONT_HEADER = Font(name="Arial", size=10, bold=True)
    FONT_TITLE = Font(name="Arial", size=12, bold=True)

    # Border class attributes
    BORDER_THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Alignment class attributes
    ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGNMENT_CENTER_WRAP = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGNMENT_RIGHT = Alignment(horizontal="right", vertical="center")

    @classmethod
    def apply_style_to_merged_cells(
        cls,
        sheet: Worksheet,
        start_col: str,
        start_row: int,
        end_col: str,
        end_row: int,
        font: Font | None = None,
        alignment: Alignment | None = None,
        border: Border | None = None,
    ) -> None:
        """Apply consistent styling to all cells in a merged range.

        Args:
            start_col: Starting column letter (e.g., 'A')
            start_row: Starting row number
            end_col: Ending column letter (e.g., 'E')
            end_row: Ending row number
            font: Font style to apply (defaults to FONT_NORMAL)
            alignment: Alignment style to apply (defaults to ALIGNMENT_CENTER)
            border: Optional border style to apply
        """
        font = font or cls.FONT_NORMAL
        alignment = alignment or cls.ALIGNMENT_CENTER

        start_col_idx: int = ord(start_col) - ord("A") + 1
        end_col_idx: int = ord(end_col) - ord("A") + 1

        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter: str = get_column_letter(col_idx)
                cell = sheet[f"{col_letter}{row}"]
                cell.font = font
                cell.alignment = alignment
                if border:
                    cell.border = border

    @classmethod
    def apply_style_to_cell(
        cls,
        sheet: Worksheet,
        col: str,
        row: int,
        font: Font | None = None,
        alignment: Alignment | None = None,
        border: Border | None = None,
        value: Any = None,
    ) -> None:
        """Apply styling to a single cell.

        Args:
            col: Column letter (e.g., 'A')
            row: Row number
            font: Font style to apply (defaults to FONT_NORMAL)
            alignment: Alignment style to apply (defaults to ALIGNMENT_CENTER)
            border: Optional border style to apply
            value: Optional value to set in the cell
        """
        font = font or cls.FONT_NORMAL
        alignment = alignment or cls.ALIGNMENT_CENTER

        cell = sheet[f"{col}{row}"]
        if value is not None:
            cell.value = value
        cell.font = font
        cell.alignment = alignment
        if border:
            cell.border = border

    @classmethod
    def merge_and_style_cells(
        cls,
        sheet: Worksheet,
        start_col: str,
        start_row: int,
        end_col: str,
        end_row: int,
        value: Any = None,
        font: Font | None = None,
        alignment: Alignment | None = None,
        border: Border | None = None,
    ) -> None:
        """Merge cells and apply styling in one operation.

        Args:
            start_col: Starting column letter (e.g., 'A')
            start_row: Starting row number
            end_col: Ending column letter (e.g., 'E')
            end_row: Ending row number
            value: Value to set in the merged cell
            font: Font style to apply (defaults to FONT_NORMAL)
            alignment: Alignment style to apply (defaults to ALIGNMENT_CENTER)
            border: Optional border style to apply
        """
        # Set the value in the top-left cell before merging
        if value is not None:
            sheet[f"{start_col}{start_row}"] = value

        # Merge the cells
        sheet.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

        # Apply styling to all cells in the merged range
        cls.apply_style_to_merged_cells(
            sheet, start_col, start_row, end_col, end_row, font, alignment, border
        )

    @classmethod
    def set_column_widths(cls, sheet: Worksheet, column_widths: dict[str, int]) -> None:
        """Set column widths for the worksheet.

        Args:
            column_widths: Dictionary mapping column letters to their widths
        """
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

    @classmethod
    def setup_page_layout(
        cls,
        sheet: Worksheet,
        orientation: Literal["portrait", "landscape"] = "portrait",
        fit_to_width: bool = True,
        fit_to_height: bool = False,
    ) -> None:
        """Setup basic page layout for the worksheet.

        Args:
            orientation: Page orientation ("portrait" or "landscape")
            fit_to_width: Whether to fit content to page width
            fit_to_height: Whether to fit content to page height
        """
        sheet.page_setup.orientation = orientation
        if fit_to_width:
            sheet.page_setup.fitToWidth = 1
        if fit_to_height:
            sheet.page_setup.fitToHeight = 1
        if not fit_to_height:
            sheet.page_setup.fitToHeight = 0

    @classmethod
    def setup_page_margins(
        cls,
        sheet: Worksheet,
        left: float = 0.25,
        right: float = 0.25,
        top: float = 0.75,
        bottom: float = 0.75,
    ) -> None:
        """Setup page margins for the worksheet.

        Args:
            left: Left margin in inches
            right: Right margin in inches
            top: Top margin in inches
            bottom: Bottom margin in inches
        """
        sheet.page_margins.left = left
        sheet.page_margins.right = right
        sheet.page_margins.top = top
        sheet.page_margins.bottom = bottom

    @classmethod
    def apply_data_row_styling(
        cls,
        sheet: Worksheet,
        row: int,
        data: list[tuple[str, Any]],
        font: Font | None = None,
        alignment: Alignment | None = None,
        border: Border | None = None,
    ) -> None:
        """Apply styling to multiple cells in a data row.

        Args:
            row: Row number to apply styling to
            data: List of tuples containing (column, value) pairs
            font: Font style to apply (defaults to FONT_NORMAL)
            alignment: Alignment style to apply (defaults to ALIGNMENT_CENTER)
            border: Optional border style to apply
        """
        font = font or cls.FONT_NORMAL
        alignment = alignment or cls.ALIGNMENT_CENTER

        for col, value in data:
            cls.apply_style_to_cell(sheet, col, row, font, alignment, border, value)

    @classmethod
    def create_custom_alignment(
        cls,
        sheet: Worksheet,
        horizontal: str = "center",
        vertical: str = "center",
        wrap_text: bool = False,
    ) -> Alignment:
        """Create a custom alignment object.

        Args:
            horizontal: Horizontal alignment ("left", "center", "right")
            vertical: Vertical alignment ("top", "center", "bottom")
            wrap_text: Whether to wrap text

        Returns:
            Alignment object with specified properties
        """
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)

    @classmethod
    def create_custom_font(
        cls,
        sheet: Worksheet,
        name: str = "Arial",
        size: int = 9,
        bold: bool = False,
        italic: bool = False,
    ) -> Font:
        """Create a custom font object.

        Args:
            name: Font name
            size: Font size
            bold: Whether font is bold
            italic: Whether font is italic

        Returns:
            Font object with specified properties
        """
        return Font(name=name, size=size, bold=bold, italic=italic)
