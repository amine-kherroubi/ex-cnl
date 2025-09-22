from __future__ import annotations
from datetime import date
from pathlib import Path
from typing import Any

# Imports tiers
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Imports de l'application locale
from app.core.domain.models.programme import Programme
from app.core.domain.models.report_context import ReportContext
from app.core.domain.models.report_specification import ReportSpecification
from app.core.domain.predefined_objects.programmes import (
    RURAL_HOUSING_PROGRAMMES,
    get_programmes_dataframe,
)
from app.core.domain.predefined_objects.dairas_et_communes import (
    get_dairas_communes_dataframe,
)
from app.core.infrastructure.data.data_repository import DataRepository
from app.core.infrastructure.file_io.file_io_service import FileIOService
from app.core.services.report_generation.base.report_generator import ReportGenerator


# Font and style constants for consistency
FONT_NORMAL = Font(name="Arial", size=9, bold=False)
FONT_BOLD = Font(name="Arial", size=9, bold=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True)
FONT_TITLE = Font(name="Arial", size=12, bold=True)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")
ALIGNMENT_CENTER_WRAP = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)


class SituationFinanciereGenerator(ReportGenerator):
    __slots__ = ("_current_row", "_target_programme", "_totals")

    def __init__(
        self,
        file_io_service: FileIOService,
        data_repository: DataRepository,
        report_specification: ReportSpecification,
        report_context: ReportContext,
    ) -> None:
        super().__init__(
            file_io_service, data_repository, report_specification, report_context
        )
        self._current_row: int = 1
        self._target_programme: Programme | None = None

    def configure(self, **kwargs: Any) -> None:
        """Set the target programme for the financial situation report."""
        programme_name: str | None = kwargs.get("target_programme")
        if programme_name is None:
            raise ValueError("additional parameter 'target_programme' is required")

        self._logger.debug(f"Setting target programme: {programme_name}")

        # Find programme in the list
        for programme in RURAL_HOUSING_PROGRAMMES:
            if programme.name == programme_name:
                self._target_programme = programme
                self._logger.info(
                    f"Target programme set: {programme.name} "
                    f"(aid_value: {programme.aid_value}, consistance: {programme.consistance})"
                )
                return

        # Programme not found
        available_programmes: list[str] = [p.name for p in RURAL_HOUSING_PROGRAMMES]
        error_msg: str = (
            f"Programme '{programme_name}' not found. "
            f"Available programmes: {available_programmes}"
        )
        self._logger.error(error_msg)
        raise ValueError(error_msg)

    def _verify_target_programme(self) -> None:
        """Verify that a target programme has been set."""
        if self._target_programme is None:
            error_msg: str = (
                "No target programme set. Please call set_target_programme() "
                "before generating the report."
            )
            self._logger.error(error_msg)
            raise RuntimeError(error_msg)

        self._logger.debug(f"Target programme verified: {self._target_programme.name}")

    def generate(
        self,
        source_file_paths: dict[str, Path],
        output_directory_path: Path,
    ) -> Path:
        """Generate the financial situation report."""
        self._verify_target_programme()
        return super().generate(source_file_paths, output_directory_path)

    def _create_predefined_tables(self) -> None:
        """Create reference tables needed for the report."""
        self._logger.debug("Creating reference tables")

        # Create programmes table
        try:
            self._logger.debug("Creating reference table 'programmes'")
            df: pd.DataFrame = get_programmes_dataframe()
            self._data_repository.create_table_from_dataframe("programmes", df)
            rows, cols = df.shape
            self._logger.info(
                f"Reference table 'programmes' created: {rows} rows and {cols} columns"
            )
        except Exception as error:
            self._logger.exception(
                f"Failed to create reference table 'programmes': {error}"
            )
            raise

        # Create dairas_communes table
        try:
            self._logger.debug("Creating reference table 'dairas_communes'")
            df: pd.DataFrame = get_dairas_communes_dataframe()
            self._data_repository.create_table_from_dataframe("dairas_communes", df)
            rows, cols = df.shape
            self._logger.info(
                f"Reference table 'dairas_communes' created: {rows} rows and {cols} columns"
            )
        except Exception as error:
            self._logger.exception(
                f"Failed to create reference table 'dairas_communes': {error}"
            )
            raise

    def _format_query_with_context(self, query_template: str) -> str:
        """Format query template with current report context and target programme."""
        self._logger.debug("Formatting query template with report context")
        assert self._target_programme is not None

        formatted_query: str = query_template

        formatted_query = (
            formatted_query.replace("{programme}", f"'{self._target_programme.name}'")
            .replace("{aid_value}", str(self._target_programme.aid_value))
            .replace("{year}", str(self._report_context.year))
            .replace("{month}", str(self._report_context.month.number))
        )

        self._logger.debug(
            f"Placeholders replaced with: month={self._report_context.month.number}, "
            f"year={self._report_context.year}, "
            f"programme={self._target_programme.name}, "
            f"aid_value={self._target_programme.aid_value}"
        )

        self._logger.debug("Query formatting completed")
        return formatted_query

    def _apply_style_to_merged_cells(
        self,
        sheet: Worksheet,
        start_col: str,
        start_row: int,
        end_col: str,
        end_row: int,
        font: Font = FONT_NORMAL,
        alignment: Alignment = ALIGNMENT_CENTER,
        border: Border | None = None,
    ) -> None:
        """Apply consistent styling to all cells in a merged range."""
        start_col_idx: int = ord(start_col) - ord("A") + 1
        end_col_idx: int = ord(end_col) - ord("A") + 1

        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter: str = get_column_letter(col_idx)
                cell = sheet[f"{col_letter}{row}"]
                cell.font = font
                cell.alignment = alignment
                if border:
                    cell.border = border

    def _add_content(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        """Add all content to the worksheet."""
        self._add_header(sheet)
        self._add_table_headers(sheet)
        self._add_data_rows(sheet, query_results)
        self._add_totals_row(sheet, query_results)

    def _add_header(self, sheet: Worksheet) -> None:
        """Add the report header section."""
        self._logger.debug("Adding report header")

        # Main title
        sheet[f"A{self._current_row}"] = (
            f"Situation financière du programme '{self._target_programme.name}' par daira et par commune"  # type: ignore
        )
        sheet.merge_cells(f"A{self._current_row}:T{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "A",
            self._current_row,
            "T",
            self._current_row,
            font=FONT_TITLE,
            alignment=ALIGNMENT_CENTER,
        )

        self._current_row += 1

        # Report date
        sheet[f"A{self._current_row}"] = (
            f"Arrêté au {self._report_context.reporting_date.strftime('%d/%m/%Y')}"
        )
        sheet.merge_cells(f"A{self._current_row}:T{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "A",
            self._current_row,
            "T",
            self._current_row,
            font=FONT_HEADER,
            alignment=ALIGNMENT_CENTER,
        )

        self._current_row += 1

        # Wilaya
        sheet[f"A{self._current_row}"] = f"DL de {self._report_context.wilaya.value}"
        sheet[f"A{self._current_row}"].font = FONT_HEADER
        sheet[f"A{self._current_row}"].alignment = ALIGNMENT_CENTER

        self._current_row += 2

        self._logger.info("Report header added successfully")

    def _add_table_headers(self, sheet: Worksheet) -> None:
        """Add the complex table headers with merged cells."""
        self._logger.debug("Adding table headers")

        # First level headers (engagement types)
        sheet[f"F{self._current_row}"] = "Engagement par la BNH"
        sheet.merge_cells(f"F{self._current_row}:G{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "F",
            self._current_row,
            "G",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
        )

        sheet[f"H{self._current_row}"] = (
            "Engagement par le MHUV (décision d'inscription)"
        )
        sheet.merge_cells(f"H{self._current_row}:I{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "H",
            self._current_row,
            "I",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
        )

        self._current_row += 1
        header_start_row: int = self._current_row
        header_end_row: int = self._current_row + 3

        # Main column headers with vertical spans
        main_headers: list[tuple[str, str]] = [
            ("A", "Programme"),
            ("B", "Daira"),
            ("C", "Commune"),
            ("D", "Aides notifiées (1)"),
            ("E", "Montants notifiés"),
            ("F", "Aides inscrites"),
            ("G", "Montants inscrits"),
            ("H", "Aides inscrites (2)"),
            ("I", "Montants inscrits (3)"),
            ("R", "Cumul (6) = (4) + (5)"),
            ("S", "Solde sur engagement (3) - (6)"),
            ("T", "Reste à inscrire (1) - (2)"),
        ]

        for col, title in main_headers:
            sheet[f"{col}{self._current_row}"] = title
            sheet.merge_cells(f"{col}{header_start_row}:{col}{header_end_row}")
            self._apply_style_to_merged_cells(
                sheet,
                col,
                header_start_row,
                col,
                header_end_row,
                font=FONT_BOLD,
                alignment=ALIGNMENT_CENTER_WRAP,
                border=BORDER_THIN,
            )

        # Consommations header
        sheet[f"J{self._current_row}"] = "Consommations"
        sheet.merge_cells(f"J{self._current_row}:Q{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "J",
            self._current_row,
            "Q",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
            border=BORDER_THIN,
        )

        self._current_row += 1

        # Second level headers (time periods)
        sheet[f"J{self._current_row}"] = (
            f"Cumuls au 31/12/{self._report_context.year - 1}"
        )
        sheet.merge_cells(f"J{self._current_row}:M{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "J",
            self._current_row,
            "M",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
            border=BORDER_THIN,
        )

        end_day: int = (
            self._report_context.month.last_day(self._report_context.year)
            if not self._report_context.month.is_current
            else date.today().day
        )

        sheet[f"N{self._current_row}"] = (
            f"Du 1 janvier {self._report_context.year} au {end_day} {self._report_context.month.value}"
        )
        sheet.merge_cells(f"N{self._current_row}:Q{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "N",
            self._current_row,
            "Q",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
            border=BORDER_THIN,
        )

        self._current_row += 1

        # Third level headers (aid categories)
        sheet[f"J{self._current_row}"] = "Aides"
        sheet.merge_cells(f"J{self._current_row}:L{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "J",
            self._current_row,
            "L",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
            border=BORDER_THIN,
        )

        sheet[f"M{self._current_row}"] = "Montant (4)"
        sheet[f"M{self._current_row}"].font = FONT_BOLD
        sheet[f"M{self._current_row}"].alignment = ALIGNMENT_CENTER_WRAP
        sheet[f"M{self._current_row}"].border = BORDER_THIN

        sheet[f"N{self._current_row}"] = "Aides"
        sheet.merge_cells(f"N{self._current_row}:P{self._current_row}")
        self._apply_style_to_merged_cells(
            sheet,
            "N",
            self._current_row,
            "P",
            self._current_row,
            font=FONT_BOLD,
            alignment=ALIGNMENT_CENTER_WRAP,
            border=BORDER_THIN,
        )

        sheet[f"Q{self._current_row}"] = "Montant (5)"
        sheet[f"Q{self._current_row}"].font = FONT_BOLD
        sheet[f"Q{self._current_row}"].alignment = ALIGNMENT_CENTER_WRAP
        sheet[f"Q{self._current_row}"].border = BORDER_THIN

        self._current_row += 1

        # Fourth level headers (tranche types)
        tranche_headers: list[tuple[str, str]] = [
            ("J", "T1"),
            ("K", "T2"),
            ("L", "T3"),
            ("N", "T1"),
            ("O", "T2"),
            ("P", "T3"),
        ]

        for col, title in tranche_headers:
            sheet[f"{col}{self._current_row}"] = title
            sheet[f"{col}{self._current_row}"].font = FONT_BOLD
            sheet[f"{col}{self._current_row}"].alignment = ALIGNMENT_CENTER_WRAP
            sheet[f"{col}{self._current_row}"].border = BORDER_THIN

        self._current_row += 1
        self._logger.info("Table headers added successfully")

    def _add_data_rows(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        """Add data rows to the table."""
        self._logger.debug("Adding data rows")

        # Get dairas and communes data
        dairas_communes_df: pd.DataFrame = get_dairas_communes_dataframe()

        # Create lookup dictionaries from query results
        data_dicts: dict[str, dict] = self._create_lookup_dictionaries(query_results)

        # Track totals
        totals: dict[str, int] = {
            "aides_inscrites": 0,
            "montants_inscrits": 0,
            "cumul_precedent_t1": 0,
            "cumul_precedent_t2": 0,
            "cumul_precedent_t3": 0,
            "cumul_precedent_montant": 0,
            "annee_actuelle_t1": 0,
            "annee_actuelle_t2": 0,
            "annee_actuelle_t3": 0,
            "annee_actuelle_montant": 0,
            "cumul_total": 0,
        }

        # Add data rows for each daira/commune combination
        for i, (_, row) in enumerate(dairas_communes_df.iterrows()):
            daira: str = row["Daira"]
            commune: str = row["Commune"]
            current_row: int = self._current_row + i

            self._add_single_data_row(
                sheet, current_row, daira, commune, data_dicts, totals
            )

        self._current_row += len(dairas_communes_df)

        # Store totals for use in totals row
        self._totals = totals

        self._logger.info(f"Added {len(dairas_communes_df)} data rows successfully")

    def _create_lookup_dictionaries(
        self, query_results: dict[str, pd.DataFrame]
    ) -> dict[str, dict]:
        """Create lookup dictionaries from query results."""
        data_dicts: dict[str, dict] = {
            "aides_inscrites": {},
            "cumul_precedent": {},
            "annee_actuelle": {},
        }

        # Aides inscrites
        if "nb_aides_et_montants_inscrits_par_daira_et_commune" in query_results:
            df = query_results["nb_aides_et_montants_inscrits_par_daira_et_commune"]
            for _, row in df.iterrows():
                key = (row["Daira du projet"], row["Commune du projet"])
                data_dicts["aides_inscrites"][key] = (
                    row["nb_aides_inscrites"],
                    row["montant_inscrits"],
                )

        # Cumul précédent
        if "consommations_cumulees_fin_annee_precedente" in query_results:
            df = query_results["consommations_cumulees_fin_annee_precedente"]
            for _, row in df.iterrows():
                key = (row["Daira"], row["Commune de projet"])
                data_dicts["cumul_precedent"][key] = (
                    row["t_1"],
                    row["t_2"],
                    row["t_3"],
                    row["montant"],
                )

        # Année actuelle
        if "consommations_annee_actuelle_jusqua_mois_actuel" in query_results:
            df = query_results["consommations_annee_actuelle_jusqua_mois_actuel"]
            for _, row in df.iterrows():
                key = (row["Daira"], row["Commune de projet"])
                data_dicts["annee_actuelle"][key] = (
                    row["t_1"],
                    row["t_2"],
                    row["t_3"],
                    row["montant"],
                )

        return data_dicts

    def _add_single_data_row(
        self,
        sheet: Worksheet,
        row: int,
        daira: str,
        commune: str,
        data_dicts: dict[str, dict],
        totals: dict[str, int],
    ) -> None:
        """Add a single data row to the table."""
        key: tuple[str, str] = (daira, commune)

        # Basic columns
        basic_values: list[tuple[str, Any]] = [
            ("A", self._target_programme.name),  # type: ignore
            ("B", daira),
            ("C", commune),
            ("D", "-"),
            ("E", "-"),
        ]

        # Aides inscrites (F & G)
        if key in data_dicts["aides_inscrites"]:
            aides, montants = data_dicts["aides_inscrites"][key]
            basic_values.extend(
                [
                    ("F", aides if aides > 0 else "-"),
                    ("G", montants if montants > 0 else "-"),
                ]
            )
            totals["aides_inscrites"] += aides
            totals["montants_inscrits"] += montants
        else:
            basic_values.extend([("F", "-"), ("G", "-")])

        basic_values.extend([("H", "-"), ("I", "-")])

        # Apply basic values
        for col, value in basic_values:
            cell = sheet[f"{col}{row}"]
            cell.value = value
            cell.font = FONT_NORMAL
            cell.alignment = ALIGNMENT_CENTER
            cell.border = BORDER_THIN

        # Previous year cumulative (J, K, L, M)
        self._add_cumulative_data(
            sheet,
            row,
            key,
            data_dicts["cumul_precedent"],
            ["J", "K", "L", "M"],
            totals,
            [
                "cumul_precedent_t1",
                "cumul_precedent_t2",
                "cumul_precedent_t3",
                "cumul_precedent_montant",
            ],
        )

        # Current year (N, O, P, Q)
        self._add_cumulative_data(
            sheet,
            row,
            key,
            data_dicts["annee_actuelle"],
            ["N", "O", "P", "Q"],
            totals,
            [
                "annee_actuelle_t1",
                "annee_actuelle_t2",
                "annee_actuelle_t3",
                "annee_actuelle_montant",
            ],
        )

        # Calculate and add cumul total (R)
        cumul_total = self._calculate_cumul_total(key, data_dicts)
        cell = sheet[f"R{row}"]
        cell.value = cumul_total if cumul_total > 0 else "-"
        cell.font = FONT_NORMAL
        cell.alignment = ALIGNMENT_CENTER
        cell.border = BORDER_THIN
        totals["cumul_total"] += cumul_total

        # Final columns (S, T)
        for col in ["S", "T"]:
            cell = sheet[f"{col}{row}"]
            cell.value = "-"
            cell.font = FONT_NORMAL
            cell.alignment = ALIGNMENT_CENTER
            cell.border = BORDER_THIN

    def _add_cumulative_data(
        self,
        sheet: Worksheet,
        row: int,
        key: tuple[str, str],
        data_dict: dict,
        columns: list[str],
        totals: dict[str, int],
        total_keys: list[str],
    ) -> None:
        """Add cumulative data to specified columns."""
        if key in data_dict:
            values = data_dict[key]
            for i, (col, value, total_key) in enumerate(
                zip(columns, values, total_keys)
            ):
                cell = sheet[f"{col}{row}"]
                cell.value = value if value > 0 else "-"
                cell.font = FONT_NORMAL
                cell.alignment = ALIGNMENT_CENTER
                cell.border = BORDER_THIN
                totals[total_key] += value
        else:
            for col in columns:
                cell = sheet[f"{col}{row}"]
                cell.value = "-"
                cell.font = FONT_NORMAL
                cell.alignment = ALIGNMENT_CENTER
                cell.border = BORDER_THIN

    def _calculate_cumul_total(
        self, key: tuple[str, str], data_dicts: dict[str, dict]
    ) -> int:
        """Calculate the cumulative total for a given key."""
        total: int = 0
        if key in data_dicts["cumul_precedent"]:
            total += data_dicts["cumul_precedent"][key][3]
        if key in data_dicts["annee_actuelle"]:
            total += data_dicts["annee_actuelle"][key][3]
        return total

    def _add_totals_row(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        """Add the totals row at the bottom of the table."""
        self._logger.debug("Adding totals row")

        totals: dict[str, int] = getattr(self, "_totals", {})

        # Define all total values
        total_values: list[tuple[str, Any]] = [
            ("A", "Total général"),
            ("B", ""),
            ("C", ""),
            ("D", "-"),
            ("E", "-"),
            ("F", totals.get("aides_inscrites", 0)),
            ("G", totals.get("montants_inscrits", 0)),
            ("H", "-"),
            ("I", "-"),
            ("J", totals.get("cumul_precedent_t1", 0)),
            ("K", totals.get("cumul_precedent_t2", 0)),
            ("L", totals.get("cumul_precedent_t3", 0)),
            ("M", totals.get("cumul_precedent_montant", 0)),
            ("N", totals.get("annee_actuelle_t1", 0)),
            ("O", totals.get("annee_actuelle_t2", 0)),
            ("P", totals.get("annee_actuelle_t3", 0)),
            ("Q", totals.get("annee_actuelle_montant", 0)),
            ("R", totals.get("cumul_total", 0)),
            ("S", "-"),
            ("T", "-"),
        ]

        # Apply values and formatting
        for col, value in total_values:
            cell = sheet[f"{col}{self._current_row}"]
            cell.value = value
            cell.font = FONT_BOLD
            cell.alignment = ALIGNMENT_CENTER
            cell.border = BORDER_THIN

        self._logger.info("Totals row added successfully")
        self._current_row += 1

    def _finalize_formatting(self, sheet: Worksheet) -> None:
        """Apply final formatting to the worksheet."""
        self._logger.debug("Applying final formatting")

        # Set column widths for better readability
        column_widths: dict[str, int] = {
            "A": 30,  # Programme
            "B": 20,  # Daira
            "C": 25,  # Commune
            "D": 12,  # Aides notifiées
            "E": 15,  # Montants notifiés
            "F": 12,  # Aides inscrites
            "G": 15,  # Montants inscrits
            "H": 12,  # Aides inscrites (2)
            "I": 15,  # Montants inscrits (3)
            "J": 8,  # T1 (prev)
            "K": 8,  # T2 (prev)
            "L": 8,  # T3 (prev)
            "M": 15,  # Montant (4)
            "N": 8,  # T1 (curr)
            "O": 8,  # T2 (curr)
            "P": 8,  # T3 (curr)
            "Q": 15,  # Montant (5)
            "R": 15,  # Cumul
            "S": 20,  # Solde
            "T": 20,  # Reste
        }

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # Page setup for landscape orientation
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        # Set print area and margins
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.75
        sheet.page_margins.bottom = 0.75

        self._logger.info("Final formatting completed successfully")
