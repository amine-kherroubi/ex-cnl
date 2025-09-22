from __future__ import annotations

# Imports de la bibliothèque standard
from datetime import date
from typing import Any

# Imports tiers
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, Side

# Imports de l'application locale
from app.core.domain.models.report_context import ReportContext
from app.core.domain.models.report_specification import ReportSpecification
from app.core.domain.predefined_objects.programmes import get_programmes_dataframe
from app.core.infrastructure.data.data_repository import DataRepository
from app.core.infrastructure.file_io.file_io_service import FileIOService
from app.core.services.report_generation.base.report_generator import ReportGenerator


class ActiviteMensuelleGenerator(ReportGenerator):
    __slots__ = ("_current_row",)

    def __init__(
        self,
        file_io_service: FileIOService,
        data_repository: DataRepository,
        report_specification: ReportSpecification,
        report_context: ReportContext,
    ) -> None:
        super().__init__(
            file_io_service, data_repository, report_specification, report_context
        )
        self._current_row: int = 1

    def configure(self, **kwargs: Any) -> None:
        """Configure the generator with additional parameters.

        This generator doesn't need additional configuration,
        but overrides the method to avoid NotImplementedError.
        """
        if kwargs:
            self._logger.debug(
                f"Ignoring unused configuration parameters: {list(kwargs.keys())}"
            )

    def _create_predefined_tables(self) -> None:
        self._logger.debug("Creating reference tables")
        try:
            self._logger.debug(f"Creating reference table 'programmes'")

            df: pd.DataFrame = get_programmes_dataframe()
            self._data_repository.create_table_from_dataframe("programmes", df)

            rows, cols = df.shape
            self._logger.info(
                f"Reference table 'programmes' created: {rows} rows and {cols} columns"
            )
            self._logger.debug(f"Columns for 'programmes': {list(df.columns)}")

        except Exception as error:
            self._logger.exception(
                f"Failed to create reference table 'programmes': {error}"
            )
            raise

    def _format_query_with_context(self, query_template: str) -> str:
        self._logger.debug("Formatting query template with report context")

        formatted_query: str = query_template

        formatted_query = formatted_query.replace(
            "{month}", str(self._report_context.month.number)
        ).replace("{year}", str(self._report_context.year))

        self._logger.debug(
            f"Placeholders replaced with: month={self._report_context.month.number}, year={self._report_context.year}"
        )

        self._logger.debug("Query formatting completed")
        return formatted_query

    def _add_content(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        self._add_first_table_header(sheet)
        self._add_first_table(sheet, query_results)
        self._add_second_table_header(sheet)
        self._add_second_table(sheet, query_results)
        self._add_footer(sheet)

    def _add_first_table_header(self, sheet: Worksheet) -> None:
        self._logger.debug("Ajout de l'en-tête du report")

        # Titre
        sheet[f"A{self._current_row}"] = "Habitat rural"
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self._logger.debug("Titre principal ajouté : Habitat rural")

        self._current_row += 1

        # Wilaya
        wilaya_text = f"Wilaya de {self._report_context.wilaya.value}"
        sheet[f"A{self._current_row}"] = wilaya_text
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        self._logger.debug(f"Wilaya ajoutée : {wilaya_text}")

        self._current_row += 1

        # Titre principal
        sheet[f"A{self._current_row}"] = (
            "Activité mensuelle par programme (à renseigner par la BNH, ex-CNL)"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self._logger.debug("Titre du report ajouté")

        self._current_row += 1

        # Mois
        month_text: str = (
            f"Mois de {self._report_context.month} {self._report_context.year}"  # type: ignore
        )
        sheet[f"A{self._current_row}"] = month_text
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self._logger.debug(f"Mois et année ajoutés : {month_text}")

        self._logger.info("En-tête du report terminé avec succès")

        self._current_row += 2

    def _add_first_table(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        self._logger.debug("Ajout du tableau principal de données")
        self._logger.debug(
            f"Résultats de requêtes disponibles : {list(query_results.keys())}"
        )

        # La cellule Programme s'étend sur 3 lignes
        sheet[f"A{self._current_row}"] = "Programme"
        sheet.merge_cells(f"A{self._current_row}:A{self._current_row + 2}")
        for cell in [sheet[f"A{self._current_row + i}"] for i in range(3)]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        sheet[f"B{self._current_row}"] = (
            f"État d'exécution des tranches financières durant le mois de "
            f"{self._report_context.month} {self._report_context.year}"  # type: ignore
        )
        sheet.merge_cells(f"B{self._current_row}:E{self._current_row}")
        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.debug("Légende du tableau ajoutée")

        self._current_row += 1

        # En-têtes
        self._logger.debug(
            f"Ajout des en-têtes de colonnes à la ligne {self._current_row}"
        )
        sheet[f"B{self._current_row}"] = (
            "Livraisons (libération de la dernière tranche)"
        )
        sheet.merge_cells(f"B{self._current_row}:C{self._current_row}")
        sheet[f"D{self._current_row}"] = (
            "Lancements (libération de la première tranche)"
        )
        sheet.merge_cells(f"D{self._current_row}:E{self._current_row}")

        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._current_row += 1

        # Sous-en-têtes
        sheet[f"B{self._current_row}"] = (
            f"{self._report_context.month.capitalize()} {self._report_context.year}"
        )

        end_day: int = (
            self._report_context.month.last_day(self._report_context.year)
            if not self._report_context.month.is_current
            else date.today().day
        )

        sheet[f"C{self._current_row}"] = (
            f"Cumul du 1er janvier au {end_day}"
            f"{self._report_context.month} {self._report_context.year}"
        )
        sheet[f"D{self._current_row}"] = (
            f"{self._report_context.month.capitalize()} {self._report_context.year}"
        )
        sheet[f"E{self._current_row}"] = (
            f"Cumul du 1er janvier au {end_day}"
            f"{self._report_context.month} {self._report_context.year}"
        )

        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.debug("Sous-en-têtes ajoutés avec les plages de dates")

        self._current_row += 1

        # Ajouter les données des résultats de requêtes
        self._logger.debug(
            f"Début des lignes de données à la ligne {self._current_row}"
        )

        # Obtenir tous les programmes avec l'ordre approprié
        programmes: list[str] = []
        if "programmes" in query_results:
            programmes: list[str] = query_results["programmes"]["programme"].tolist()
            self._logger.info(
                f"Trouvé {len(programmes)} programmes (pré-triés par année) : {programmes}"
            )
        else:
            self._logger.warning("Aucun résultat de requête 'programmes' trouvé")

        # Créer des dictionnaires de recherche pour chaque métrique
        self._logger.debug(
            "Création des dictionnaires de recherche à partir des résultats de requêtes"
        )

        lancements_mois_dict: dict[str, int] = {}
        if "lancements_mois" in query_results:
            df_lm: pd.DataFrame = query_results["lancements_mois"]
            lancements_mois_dict = dict(zip(df_lm["programme"], df_lm["count"]))
            self._logger.debug(
                f"Données lancements mois : {len(lancements_mois_dict)} programmes"
            )

        lancements_cumul_annee_dict: dict[str, int] = {}
        if "lancements_cumul_annee" in query_results:
            df_ly: pd.DataFrame = query_results["lancements_cumul_annee"]
            lancements_cumul_annee_dict = dict(zip(df_ly["programme"], df_ly["count"]))
            self._logger.debug(
                f"Données lancements cumulé annuel : {len(lancements_cumul_annee_dict)} programmes"
            )

        livraisons_mois_dict: dict[str, int] = {}
        if "livraisons_mois" in query_results:
            df_livm: pd.DataFrame = query_results["livraisons_mois"]
            livraisons_mois_dict = dict(zip(df_livm["programme"], df_livm["count"]))
            self._logger.debug(
                f"Données livraisons mois : {len(livraisons_mois_dict)} programmes"
            )

        livraisons_cumul_annee_dict: dict[str, int] = {}
        if "livraisons_cumul_annee" in query_results:
            df_livy: pd.DataFrame = query_results["livraisons_cumul_annee"]
            livraisons_cumul_annee_dict = dict(
                zip(df_livy["programme"], df_livy["count"])
            )
            self._logger.debug(
                f"Données livraisons cumulé annuel : {len(livraisons_cumul_annee_dict)} programmes"
            )

        # Calculer les totaux
        total_livraisons_mois: int = sum(livraisons_mois_dict.values())
        total_livraisons_cumul_annee: int = sum(livraisons_cumul_annee_dict.values())
        total_lancements_mois: int = sum(lancements_mois_dict.values())
        total_lancements_cumul_annee: int = sum(lancements_cumul_annee_dict.values())

        self._logger.info(
            f"Totaux calculés - Livraisons : {total_livraisons_mois} (mois), {total_livraisons_cumul_annee} (cumulé)"
        )
        self._logger.info(
            f"Totaux calculés - Lancements : {total_lancements_mois} (mois), {total_lancements_cumul_annee} (cumulé)"
        )

        # Ajouter les lignes de données pour chaque programme
        self._logger.debug(
            f"Ajout des lignes de données pour {len(programmes)} programmes"
        )
        for i, programme in enumerate(programmes):
            row: int = self._current_row + i
            self._logger.debug(
                f"Traitement du programme '{programme}' à la ligne {row}"
            )

            # Colonne A : Nom du programme
            sheet[f"A{row}"] = programme
            sheet[f"A{row}"].font = Font(name="Arial", size=9)

            # Colonne B : Livraisons (Mois)
            livraisons_mois: int = livraisons_mois_dict.get(programme, 0)
            sheet[f"B{row}"] = livraisons_mois if livraisons_mois > 0 else "-"
            sheet[f"B{row}"].font = Font(name="Arial", size=9)

            # Colonne C : Livraisons (Cumulé)
            livraisons_cumul_annee: int = livraisons_cumul_annee_dict.get(programme, 0)
            sheet[f"C{row}"] = (
                livraisons_cumul_annee if livraisons_cumul_annee > 0 else "-"
            )
            sheet[f"C{row}"].font = Font(name="Arial", size=9)

            # Colonne D : Lancements (Mois)
            lancements_mois: int = lancements_mois_dict.get(programme, 0)
            sheet[f"D{row}"] = lancements_mois if lancements_mois > 0 else "-"
            sheet[f"D{row}"].font = Font(name="Arial", size=9)

            # Colonne E : Lancements (Cumulé)
            lancements_cumul_annee: int = lancements_cumul_annee_dict.get(programme, 0)
            sheet[f"E{row}"] = (
                lancements_cumul_annee if lancements_cumul_annee > 0 else "-"
            )
            sheet[f"E{row}"].font = Font(name="Arial", size=9)

            # Ajouter les bordures
            for column in ["A", "B", "C", "D", "E"]:
                cell = sheet[f"{column}{row}"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Ajouter la ligne Total pour le premier tableau
        self._current_row += len(programmes)
        self._logger.debug(f"Ajout de la ligne Total à la ligne {self._current_row}")

        sheet[f"A{self._current_row}"] = "Total"
        sheet[f"B{self._current_row}"] = total_livraisons_mois
        sheet[f"C{self._current_row}"] = total_livraisons_cumul_annee
        sheet[f"D{self._current_row}"] = total_lancements_mois
        sheet[f"E{self._current_row}"] = total_lancements_cumul_annee

        # Formater la ligne Total
        for column in ["A", "B", "C", "D", "E"]:
            cell: Any = sheet[f"{column}{self._current_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.info(
            f"Premier tableau terminé avec {len(programmes)} programmes plus les totaux"
        )

        self._current_row += 2

    def _add_second_table_header(self, sheet: Worksheet) -> None:
        sheet[f"A{self._current_row}"] = (
            "Situation des programmes (à renseigner par la BNH, ex-CNL)"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        self._current_row += 1

        sheet[f"A{self._current_row}"] = (
            f"Arrêté le {self._report_context.reporting_date.strftime("%d/%m/%Y")}"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        self._current_row += 2

    def _add_second_table(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        self._logger.debug("Ajout du second tableau (SITUATION DES PROGRAMMES)")

        # En-têtes du second tableau
        headers: list[tuple[str, str]] = [
            ("A", "Programme"),
            ("B", "Consistance"),
            ("C", "Achevés (dernières tranches payées)"),
            ("D", "En cours"),
            (
                "E",
                "Non lancés (consistance - premières tranches payées)",
            ),
        ]

        self._logger.debug(
            f"Ajout de {len(headers)} en-têtes de colonnes pour le second tableau à la ligne {self._current_row}"
        )
        for col, title in headers:
            cell = sheet[f"{col}{self._current_row}"]
            cell.value = title
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._current_row += 1

        # Obtenir les données des résultats de requêtes
        programmes_situation: list[tuple[str, int]] = []
        if "programmes_situation" in query_results:
            df_prog = query_results["programmes_situation"]
            programmes_situation = list(
                zip(df_prog["programme"], df_prog["consistance"])
            )
            self._logger.info(
                f"Trouvé {len(programmes_situation)} programmes pour le tableau de situation"
            )
        else:
            self._logger.warning(
                "Aucun résultat de requête 'programmes_situation' trouvé"
            )

        # Créer des dictionnaires de recherche
        acheves_dict: dict[str, int] = {}
        if "acheves_derniere_tranche" in query_results:
            df_acheves = query_results["acheves_derniere_tranche"]
            acheves_dict = dict(zip(df_acheves["programme"], df_acheves["acheves"]))
            self._logger.debug(f"Données achevés : {len(acheves_dict)} programmes")

        en_cours_dict: dict[str, int] = {}
        if "en_cours_calculation" in query_results:
            df_en_cours = query_results["en_cours_calculation"]
            en_cours_dict = dict(zip(df_en_cours["programme"], df_en_cours["en_cours"]))
            self._logger.debug(f"Données en cours : {len(en_cours_dict)} programmes")

        non_lances_dict: dict[str, int] = {}
        if "non_lances_premiere_tranche" in query_results:
            df_non_lances = query_results["non_lances_premiere_tranche"]
            non_lances_dict = dict(
                zip(df_non_lances["programme"], df_non_lances["non_lances"])
            )
            self._logger.debug(
                f"Données non lancés : {len(non_lances_dict)} programmes"
            )

        # Ajouter les lignes de données
        total_consistance = 0
        total_acheves = 0
        total_en_cours = 0
        total_non_lances = 0

        for i, (programme, consistance) in enumerate(programmes_situation):
            row = self._current_row + i
            self._logger.debug(
                f"Traitement du programme '{programme}' à la ligne {row}"
            )

            # Colonne A : Nom du programme
            sheet[f"A{row}"] = programme
            sheet[f"A{row}"].font = Font(name="Arial", size=9)

            # Colonne B : Consistance
            sheet[f"B{row}"] = consistance
            sheet[f"B{row}"].font = Font(name="Arial", size=9)
            total_consistance += consistance

            # Colonne C : Achevés
            acheves = acheves_dict.get(programme, 0)
            sheet[f"C{row}"] = acheves if acheves > 0 else "-"
            sheet[f"C{row}"].font = Font(name="Arial", size=9)
            total_acheves += acheves

            # Colonne D : En cours
            en_cours = en_cours_dict.get(programme, 0)
            sheet[f"D{row}"] = en_cours if en_cours > 0 else "-"
            sheet[f"D{row}"].font = Font(name="Arial", size=9)
            total_en_cours += en_cours

            # Colonne E : Non lancés
            non_lances = non_lances_dict.get(programme, 0)
            sheet[f"E{row}"] = non_lances if non_lances > 0 else "-"
            sheet[f"E{row}"].font = Font(name="Arial", size=9)
            total_non_lances += non_lances

            # Ajouter les bordures
            for col in ["A", "B", "C", "D", "E"]:
                cell = sheet[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Ajouter la ligne Total pour le second tableau
        self._current_row += len(programmes_situation)
        self._logger.debug(
            f"Ajout de la ligne Total pour le second tableau à la ligne {self._current_row}"
        )

        sheet[f"A{self._current_row}"] = "Total général"
        sheet[f"B{self._current_row}"] = total_consistance
        sheet[f"C{self._current_row}"] = total_acheves
        sheet[f"D{self._current_row}"] = total_en_cours
        sheet[f"E{self._current_row}"] = total_non_lances

        # Formater la ligne Total
        for col in ["A", "B", "C", "D", "E"]:
            cell = sheet[f"{col}{self._current_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.info(
            f"Second tableau terminé avec {len(programmes_situation)} programmes plus les totaux"
        )
        self._current_row += 2

    def _add_footer(self, sheet: Worksheet) -> None:
        self._logger.debug("Ajout du pied de page du report")

        # Texte de pied de page gauche (A-B)
        sheet.merge_cells(f"A{self._current_row}:B{self._current_row}")
        sheet[f"A{self._current_row}"] = "Visa du directeur régional de la BNH (ex-CNL)"
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Texte de pied de page droit (D-E)
        sheet.merge_cells(f"D{self._current_row}:E{self._current_row}")
        sheet[f"D{self._current_row}"] = "Visa du directeur du logement"
        sheet[f"D{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"D{self._current_row}"].alignment = Alignment(
            horizontal="right", vertical="center"
        )

        self._logger.debug("Pied de page ajouté avec succès")

    def _finalize_formatting(self, sheet: Worksheet) -> None:
        self._logger.debug("Application du formatage final")

        column_widths: dict[str, int] = {"A": 25, "B": 18, "C": 22, "D": 18, "E": 22}
        self._logger.debug(f"Définition des largeurs de colonnes : {column_widths}")

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        self._logger.debug(
            "Orientation de page définie en portrait avec ajustement à la largeur"
        )
        self._logger.info("Formatage final terminé avec succès")
