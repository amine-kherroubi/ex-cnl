# Standard library imports
from typing import Any, Dict, List, Literal, NamedTuple, Optional, final, Final

# Third-party imports
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string


class CellData(NamedTuple):
    column: str
    row: int
    value: Any
    font: Optional[Font] = None
    alignment: Optional[Alignment] = None
    border: Optional[Border] = None


class MergeData(NamedTuple):
    start_column: str
    end_column: str
    start_row: int
    end_row: int
    value: Any = None
    font: Optional[Font] = None
    alignment: Optional[Alignment] = None
    border: Optional[Border] = None


class RowData(NamedTuple):
    number: int
    cells: List[CellData]


@final
class ExcelStylingService:
    __slots__ = ()

    FONT_NORMAL: Final[Font] = Font(name="Arial", size=8, bold=False)
    FONT_BOLD: Final[Font] = Font(name="Arial", size=8, bold=True)

    BORDER_THIN: Final[Border] = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ALIGN_CENTER: Final[Alignment] = Alignment(horizontal="center", vertical="center")
    ALIGN_CENTER_WRAP: Final[Alignment] = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ALIGN_LEFT: Final[Alignment] = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT: Final[Alignment] = Alignment(horizontal="right", vertical="center")
    ALIGN_LEFT_WRAP: Final[Alignment] = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    LETTER_TO_INDEX: Final[Dict[str, int]] = {chr(i + 65): i + 1 for i in range(26)}
    INDEX_TO_LETTER: Final[Dict[int, str]] = {i + 1: chr(i + 65) for i in range(26)}

    @classmethod
    def get_column_letter_fast(cls, index: int) -> str:
        return cls.INDEX_TO_LETTER.get(index, get_column_letter(index))

    @classmethod
    def get_column_index_fast(cls, letter: str) -> int:
        return cls.LETTER_TO_INDEX.get(letter, column_index_from_string(letter))

    @classmethod
    def style_cell(cls, sheet: Worksheet, data: CellData) -> None:
        cell: Cell = sheet[f"{data.column}{data.row}"]
        cell.value = data.value
        cell.font = data.font or cls.FONT_NORMAL
        cell.alignment = data.alignment or cls.ALIGN_CENTER
        if data.border is not None:
            cell.border = data.border

    @classmethod
    def style_row(cls, sheet: Worksheet, data: RowData) -> None:
        for cell_data in data.cells:
            cell: Cell = sheet[f"{cell_data.column}{data.number}"]
            cell.value = cell_data.value
            cell.font = cell_data.font or cls.FONT_NORMAL
            cell.alignment = cell_data.alignment or cls.ALIGN_CENTER
            if cell_data.border is not None:
                cell.border = cell_data.border

    @classmethod
    def batch_style_rows(cls, sheet: Worksheet, data_list: List[RowData]) -> None:
        for data in data_list:
            cls.style_row(sheet, data)

    @classmethod
    def style_column(cls, sheet: Worksheet, column: str, width: int) -> None:
        sheet.column_dimensions[column].width = width

    @classmethod
    def batch_style_columns(
        cls, sheet: Worksheet, column_widths: Dict[str, int]
    ) -> None:
        for column, width in column_widths.items():
            cls.style_column(sheet, column, width)

    @classmethod
    def merge_and_style_cell(cls, sheet: Worksheet, data: MergeData) -> None:
        if data.value is not None:
            anchor: Cell = sheet[f"{data.start_column}{data.start_row}"]
            anchor.value = data.value
        sheet.merge_cells(
            f"{data.start_column}{data.start_row}:{data.end_column}{data.end_row}"
        )
        anchor = sheet[f"{data.start_column}{data.start_row}"]
        anchor.font = data.font or cls.FONT_NORMAL
        anchor.alignment = data.alignment or cls.ALIGN_CENTER
        if data.border is not None:
            start_idx = cls.get_column_index_fast(data.start_column)
            end_idx = cls.get_column_index_fast(data.end_column)
            for row in range(data.start_row, data.end_row + 1):
                for col_idx in range(start_idx, end_idx + 1):
                    letter = cls.get_column_letter_fast(col_idx)
                    cell: Cell = sheet[f"{letter}{row}"]
                    cell.border = data.border

    @classmethod
    def batch_merge_and_style_cells(
        cls, sheet: Worksheet, data_list: List[MergeData]
    ) -> None:
        for data in data_list:
            cls.merge_and_style_cell(sheet, data)

    @classmethod
    def merge_and_style_cells_with_same_value(
        cls,
        sheet: Worksheet,
        column: str,
        start_row: int,
        end_row: int,
        font: Optional[Font] = None,
        alignment: Optional[Alignment] = None,
        border: Optional[Border] = None,
    ) -> None:
        base_value: Any = sheet[f"{column}{start_row}"].value
        for row in range(start_row + 1, end_row + 1):
            if sheet[f"{column}{row}"].value == base_value:
                sheet[f"{column}{row}"].value = ""
            else:
                raise ValueError(
                    f"Different values in {column}{start_row}:{column}{end_row}"
                )
        data = MergeData(
            column, column, start_row, end_row, None, font, alignment, border
        )
        cls.merge_and_style_cell(sheet, data)

    @classmethod
    def create_sum_formula(cls, column: str, start_row: int, end_row: int) -> str:
        return f"=SUM({column}{start_row}:{column}{end_row})"

    @classmethod
    def format_numbers(
        cls, sheet: Worksheet, columns: List[str], start_row: int, end_row: int
    ) -> None:
        for col in columns:
            for row in range(start_row, end_row):
                cell = sheet[f"{col}{row}"]
                cell.number_format = "#,##0"

    @classmethod
    def set_page_layout(
        cls,
        sheet: Worksheet,
        orientation: Literal["portrait", "landscape"] = "portrait",
        fit_to_width: bool = True,
        fit_to_height: bool = False,
    ) -> None:
        sheet.page_setup.orientation = orientation
        sheet.page_setup.fitToWidth = 1 if fit_to_width else 0
        sheet.page_setup.fitToHeight = 1 if fit_to_height else 0

    @classmethod
    def set_page_margins(
        cls,
        sheet: Worksheet,
        left: float = 0.25,
        right: float = 0.25,
        top: float = 0.5,
        bottom: float = 0.5,
    ) -> None:
        margins = sheet.page_margins
        margins.left = left
        margins.right = right
        margins.top = top
        margins.bottom = bottom
