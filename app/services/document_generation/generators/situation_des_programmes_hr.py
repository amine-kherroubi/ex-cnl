from __future__ import annotations
import pandas
from openpyxl.worksheet.worksheet import Worksheet
from app.services.document_generation.generator_template import DocumentGenerator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class SituationDesProgrammesHRGenerator(DocumentGenerator):
    def _add_header(self, sheet: Worksheet) -> None:
        # HABITAT RURAL
        sheet.merge_cells("A1:E1")
        sheet["C1"] = "SITUATION DES PROGRAMMES (À renseigner par la BNH (ex-CNL))"
        sheet["C1"].font = Font(name="Arial", size=9, bold=True)
        sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")

        # Date
        sheet.merge_cells("A2:E2")
        sheet["C2"] = "ARRÊTÉE AU : 20/12/2024"
        sheet["C2"].font = Font(name="Arial", size=9, bold=True)

    def _add_table(
        self, sheet: Worksheet, query_results: dict[str, pandas.DataFrame]
    ) -> None:
        start_row: int = 6

        # Caption row (part of table)
        sheet.merge_cells(f"A{start_row}:E{start_row}")
        sheet[f"C{start_row}"] = (
            "ETAT D'EXECUTION DES TRANCHES FINANCIERES DURANT LE MOIS DE JANVIER 2021"
        )
        sheet[f"C{start_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"C{start_row}"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet[f"C{start_row}"].fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )

        # Headers
        header_row: int = start_row + 2
        headers: list[tuple[str, str]] = [
            ("A", "PROGRAMMES"),
            ("B", "LIVRAISONS (libération de la dernière TR)"),
            ("C", "CUMUL LIVRAISONS"),
            ("D", "LANCEMENTS (libération de la 1ère TR)"),
            ("E", "CUMUL LANCEMENTS"),
        ]

        for col, title in headers:
            cell = sheet[f"{col}{header_row}"]
            cell.value = title
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Sub-headers
        sub_row: int = header_row + 1
        sheet[f"B{sub_row}"] = "Jan-21"
        sheet[f"C{sub_row}"] = "Cumul de JANVIER au 31 JANVIER 2021"
        sheet[f"D{sub_row}"] = "Jan-21"
        sheet[f"E{sub_row}"] = "Cumul de JANVIER au 31 JANVIER 2021"

        for col in ["B", "C", "D", "E"]:
            cell = sheet[f"{col}{sub_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    def _add_footer(self, sheet: Worksheet) -> None:
        last_row: int = sheet.max_row + 2  # leave one blank line before footer

        # Left footer text
        sheet[f"A{last_row}"] = "VISA DU DIRECTEUR REGIONAL BNH (ex-CNL)"
        sheet[f"A{last_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{last_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Right footer text
        sheet[f"C{last_row}"] = "VISA DU DIRECTEUR DU LOGEMENT"
        sheet[f"C{last_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"C{last_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

    def _finalize_formatting(self, sheet: Worksheet) -> None:
        column_widths: dict[str, int] = {"A": 25, "B": 18, "C": 22, "D": 18, "E": 22}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
