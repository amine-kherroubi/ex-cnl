from __future__ import annotations

# Third-party imports
import pandas
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Local application imports
from app.services.document_generation.generator_template import DocumentGenerator


class SituationDesProgrammesHRGenerator(DocumentGenerator):
    __slots__ = ()

    def _add_header(self, sheet: Worksheet) -> None:
        # Main title - Set value first, then merge
        sheet["A1"] = "SITUATION DES PROGRAMMES (À renseigner par la BNH (ex-CNL))"
        sheet.merge_cells("A1:E1")
        sheet["A1"].font = Font(name="Arial", size=9, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Date - Set value first, then merge
        sheet["A2"] = "ARRÊTÉE AU : 20/12/2024"
        sheet.merge_cells("A2:E2")
        sheet["A2"].font = Font(name="Arial", size=9, bold=True)
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    def _add_table(
        self, sheet: Worksheet, query_results: dict[str, pandas.DataFrame]
    ) -> None:
        start_row: int = 4

        # Caption row (part of table) - Set value first, then merge
        sheet[f"A{start_row}"] = "SITUATION DES PROGRAMMES AU 20/12/2024"
        sheet.merge_cells(f"A{start_row}:E{start_row}")
        sheet[f"A{start_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{start_row}"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        sheet[f"A{start_row}"].fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )

        # Headers
        header_row: int = start_row + 2
        headers: list[tuple[str, str]] = [
            ("A", "PROGRAMMES"),
            ("B", "LIVRAISONS (libération de la dernière TR)"),
            ("C", "CUMUL LIVRAISONS"),
            ("D", "LANCEMENTS (libération de la 1ère TR)"),
            ("E", "CUMUL LANCEMENTS"),
        ]

        for col, title in headers:
            cell = sheet[f"{col}{header_row}"]
            cell.value = title
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Sub-headers
        sub_row: int = header_row + 1
        sheet[f"B{sub_row}"] = "2024"
        sheet[f"C{sub_row}"] = "Cumul au 20/12/2024"
        sheet[f"D{sub_row}"] = "2024"
        sheet[f"E{sub_row}"] = "Cumul au 20/12/2024"

        for col in ["B", "C", "D", "E"]:
            cell = sheet[f"{col}{sub_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Add some sample data rows for demonstration
        data_start_row = sub_row + 1
        sample_programs = [
            "LOGEMENT PUBLIC LOCATIF (LPL)",
            "LOGEMENT SOCIAL PARTICIPATIF (LSP)",
            "HABITAT RURAL",
            "TOTAL",
        ]

        for i, program in enumerate(sample_programs):
            row = data_start_row + i
            sheet[f"A{row}"] = program

            # Add borders and formatting
            for col in ["A", "B", "C", "D", "E"]:
                cell = sheet[f"{col}{row}"]
                if col == "A":
                    cell.font = Font(name="Arial", size=9, bold=(program == "TOTAL"))
                else:
                    cell.value = 0  # Placeholder values
                    cell.font = Font(name="Arial", size=9)

                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                if program == "TOTAL":
                    cell.fill = PatternFill(
                        start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
                    )

    def _add_footer(self, sheet: Worksheet) -> None:
        last_row: int = sheet.max_row + 2  # leave one blank line before footer

        # Left footer text
        sheet[f"A{last_row}"] = "VISA DU DIRECTEUR REGIONAL BNH (ex-CNL)"
        sheet[f"A{last_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{last_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Right footer text
        sheet[f"C{last_row}"] = "VISA DU DIRECTEUR DU LOGEMENT"
        sheet[f"C{last_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"C{last_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

    def _finalize_formatting(self, sheet: Worksheet) -> None:
        column_widths: dict[str, int] = {"A": 25, "B": 18, "C": 22, "D": 18, "E": 22}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
