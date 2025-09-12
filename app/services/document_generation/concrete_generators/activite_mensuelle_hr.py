from __future__ import annotations

# Standard library imports
from typing import Any

# Third-party imports
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, Side

# Local application imports
from app.data.data_repository import DataRepository
from app.services.document_generation.document_generator_template import (
    DocumentGenerator,
)
from app.services.document_generation.models.document_context import DocumentContext
from app.services.document_generation.models.document_specification import (
    DocumentSpecification,
)
from app.services.io.io_service import IOService


class ActiviteMensuelleHRGenerator(DocumentGenerator):
    __slots__ = ("_current_row",)

    def __init__(
        self,
        storage_service: IOService,
        data_repository: DataRepository,
        document_specification: DocumentSpecification,
        document_context: DocumentContext,
    ) -> None:
        super().__init__(
            storage_service, data_repository, document_specification, document_context
        )
        self._current_row: int = 1

    def _add_header(self, sheet: Worksheet) -> None:
        self._logger.debug("Adding document header")

        # HABITAT RURAL - Set value first, then merge
        sheet[f"A{self._current_row}"] = "Habitat rural"
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self._logger.debug("Added main title: Habitat rural")

        self._current_row += 1

        # Wilaya
        wilaya_text = f"Wilaya de {self._document_context.wilaya.value}"
        sheet[f"A{self._current_row}"] = wilaya_text
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        self._logger.debug(f"Added wilaya: {wilaya_text}")

        # Main title - Set value first, then merge
        self._current_row += 1
        sheet[f"A{self._current_row}"] = (
            "Activité mensuelle par programme (à renseigner par la BNH, ex-CNL)"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self._logger.debug("Added document title")

        # Month - Set value first, then merge
        self._current_row += 1
        month_text: str = (
            f"Mois de {self._document_context.month.value} {self._document_context.year}"  # type: ignore
        )
        sheet[f"A{self._current_row}"] = month_text
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self._logger.debug(f"Added month and year: {month_text}")

        self._logger.info("Document header completed successfully")

        self._current_row += 2

    def _add_tables(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        self._logger.debug("Adding main data table")
        self._logger.debug(f"Available query results: {list(query_results.keys())}")

        # Programme cell spans 3 rows
        sheet[f"A{self._current_row}"] = "Programme"
        sheet.merge_cells(f"A{self._current_row}:A{self._current_row + 2}")
        for cell in [sheet[f"A{self._current_row + i}"] for i in range(3)]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        sheet[f"B{self._current_row}"] = (
            f"État d'exécution des tranches financières durant le mois de "
            f"{self._document_context.month.value} {self._document_context.year}"  # type: ignore
        )
        sheet.merge_cells(f"B{self._current_row}:E{self._current_row}")
        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.debug("Added table caption")

        # Headers
        self._logger.debug(f"Adding column headers at row {self._current_row}")
        self._current_row += 1
        sheet[f"B{self._current_row}"] = (
            "Livraisons (libération de la dernière tranche)"
        )
        sheet.merge_cells(f"B{self._current_row}:C{self._current_row}")
        sheet[f"D{self._current_row}"] = (
            "Lancements (libération de la première tranche)"
        )
        sheet.merge_cells(f"D{self._current_row}:E{self._current_row}")

        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Subheaders
        self._current_row += 1
        sheet[f"B{self._current_row}"] = (
            f"{self._document_context.month.value.capitalize()} {self._document_context.year}"  # type: ignore
        )
        sheet[f"C{self._current_row}"] = (
            f"Cumul du 1er janvier au 31 {self._document_context.month.value} {self._document_context.year}"  # type: ignore
        )
        sheet[f"D{self._current_row}"] = (
            f"{self._document_context.month.value.capitalize()} {self._document_context.year}"  # type: ignore
        )
        sheet[f"E{self._current_row}"] = (
            f"Cumul du 1er janvier au 31 {self._document_context.month.value} {self._document_context.year}"  # type: ignore
        )

        for cell in [
            sheet[f"{column}{self._current_row}"] for column in ("B", "C", "D", "E")
        ]:
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.debug("Added sub-headers with date ranges")

        # Add data from query results (existing implementation)
        self._current_row += 1
        self._logger.debug(f"Starting data rows at row {self._current_row}")

        # Get all programmes with proper ordering
        programmes: list[str] = []
        if "programmes" in query_results:
            programmes: list[str] = query_results["programmes"]["programme"].tolist()
            self._logger.info(
                f"Found {len(programmes)} programmes (pre-sorted by year): {programmes}"
            )
        else:
            self._logger.warning("No 'programmes' query result found")

        # Create lookup dictionaries for each metric
        self._logger.debug("Creating lookup dictionaries from query results")

        lancements_mois_dict: dict[str, int] = {}
        if "lancements_mois" in query_results:
            df_lm: pd.DataFrame = query_results["lancements_mois"]
            lancements_mois_dict = dict(zip(df_lm["programme"], df_lm["count"]))
            self._logger.debug(
                f"Lancements month data: {len(lancements_mois_dict)} programmes"
            )

        lancements_cumul_annee_dict: dict[str, int] = {}
        if "lancements_cumul_annee" in query_results:
            df_ly: pd.DataFrame = query_results["lancements_cumul_annee"]
            lancements_cumul_annee_dict = dict(zip(df_ly["programme"], df_ly["count"]))
            self._logger.debug(
                f"Lancements YTD data: {len(lancements_cumul_annee_dict)} programmes"
            )

        livraisons_mois_dict: dict[str, int] = {}
        if "livraisons_mois" in query_results:
            df_livm: pd.DataFrame = query_results["livraisons_mois"]
            livraisons_mois_dict = dict(zip(df_livm["programme"], df_livm["count"]))
            self._logger.debug(
                f"Livraisons month data: {len(livraisons_mois_dict)} programmes"
            )

        livraisons_cumul_annee_dict: dict[str, int] = {}
        if "livraisons_cumul_annee" in query_results:
            df_livy: pd.DataFrame = query_results["livraisons_cumul_annee"]
            livraisons_cumul_annee_dict = dict(
                zip(df_livy["programme"], df_livy["count"])
            )
            self._logger.debug(
                f"Livraisons YTD data: {len(livraisons_cumul_annee_dict)} programmes"
            )

        # Calculate totals
        total_livraisons_mois: int = sum(livraisons_mois_dict.values())
        total_livraisons_cumul_annee: int = sum(livraisons_cumul_annee_dict.values())
        total_lancements_mois: int = sum(lancements_mois_dict.values())
        total_lancements_cumul_annee: int = sum(lancements_cumul_annee_dict.values())

        self._logger.info(
            f"Calculated totals - Livraisons: {total_livraisons_mois} (month), {total_livraisons_cumul_annee} (YTD)"
        )
        self._logger.info(
            f"Calculated totals - Lancements: {total_lancements_mois} (month), {total_lancements_cumul_annee} (YTD)"
        )

        # Add data rows for each programme
        self._logger.debug(f"Adding data rows for {len(programmes)} programmes")
        for i, programme in enumerate(programmes):
            row: int = self._current_row + i
            self._logger.debug(f"Processing programme '{programme}' at row {row}")

            # Column A: Programme name
            sheet[f"A{row}"] = programme
            sheet[f"A{row}"].font = Font(name="Arial", size=9)

            # Column B: Livraisons (Month)
            livraisons_mois: int = livraisons_mois_dict.get(programme, 0)
            sheet[f"B{row}"] = livraisons_mois if livraisons_mois > 0 else "-"
            sheet[f"B{row}"].font = Font(name="Arial", size=9)

            # Column C: Livraisons (YTD)
            livraisons_cumul_annee: int = livraisons_cumul_annee_dict.get(programme, 0)
            sheet[f"C{row}"] = (
                livraisons_cumul_annee if livraisons_cumul_annee > 0 else "-"
            )
            sheet[f"C{row}"].font = Font(name="Arial", size=9)

            # Column D: Lancements (Month)
            lancements_mois: int = lancements_mois_dict.get(programme, 0)
            sheet[f"D{row}"] = lancements_mois if lancements_mois > 0 else "-"
            sheet[f"D{row}"].font = Font(name="Arial", size=9)

            # Column E: Lancements (YTD)
            lancements_cumul_annee: int = lancements_cumul_annee_dict.get(programme, 0)
            sheet[f"E{row}"] = (
                lancements_cumul_annee if lancements_cumul_annee > 0 else "-"
            )
            sheet[f"E{row}"].font = Font(name="Arial", size=9)

            # Add borders
            for column in ["A", "B", "C", "D", "E"]:
                cell = sheet[f"{column}{row}"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Add Total row for first table
        self._current_row += len(programmes)
        self._logger.debug(f"Adding Total row at row {self._current_row}")

        sheet[f"A{self._current_row}"] = "Total"
        sheet[f"B{self._current_row}"] = total_livraisons_mois
        sheet[f"C{self._current_row}"] = total_livraisons_cumul_annee
        sheet[f"D{self._current_row}"] = total_lancements_mois
        sheet[f"E{self._current_row}"] = total_lancements_cumul_annee

        # Format Total row
        for column in ["A", "B", "C", "D", "E"]:
            cell: Any = sheet[f"{column}{self._current_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.info(
            f"First table completed with {len(programmes)} programmes plus totals"
        )

        self._current_row += 2

        # Add second table header
        self._add_second_table_header(sheet)

        # Add second table
        self._add_second_table(sheet, query_results)

    def _add_second_table(
        self, sheet: Worksheet, query_results: dict[str, pd.DataFrame]
    ) -> None:
        self._logger.debug("Adding second table (SITUATION DES PROGRAMMES)")

        # Headers for second table
        headers: list[tuple[str, str]] = [
            ("A", "Programme"),
            ("B", "Consistance"),
            ("C", "Achevés (dernières tranches payées)"),
            ("D", "En cours"),
            (
                "E",
                "Non lancés (consistance - premières tranches payées)",
            ),
        ]

        self._logger.debug(
            f"Adding {len(headers)} column headers for second table at row {self._current_row}"
        )
        for col, title in headers:
            cell = sheet[f"{col}{self._current_row}"]
            cell.value = title
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._current_row += 1

        # Get data from query results
        programmes_situation: list[tuple[str, int]] = []
        if "programmes_situation" in query_results:
            df_prog = query_results["programmes_situation"]
            programmes_situation = list(
                zip(df_prog["programme"], df_prog["consistance"])
            )
            self._logger.info(
                f"Found {len(programmes_situation)} programmes for situation table"
            )
        else:
            self._logger.warning("No 'programmes_situation' query result found")

        # Create lookup dictionaries
        acheves_dict: dict[str, int] = {}
        if "acheves_derniere_tranche" in query_results:
            df_acheves = query_results["acheves_derniere_tranche"]
            acheves_dict = dict(zip(df_acheves["programme"], df_acheves["acheves"]))
            self._logger.debug(f"Achevés data: {len(acheves_dict)} programmes")

        en_cours_dict: dict[str, int] = {}
        if "en_cours_calculation" in query_results:
            df_en_cours = query_results["en_cours_calculation"]
            en_cours_dict = dict(zip(df_en_cours["programme"], df_en_cours["en_cours"]))
            self._logger.debug(f"En cours data: {len(en_cours_dict)} programmes")

        non_lances_dict: dict[str, int] = {}
        if "non_lances_premiere_tranche" in query_results:
            df_non_lances = query_results["non_lances_premiere_tranche"]
            non_lances_dict = dict(
                zip(df_non_lances["programme"], df_non_lances["non_lances"])
            )
            self._logger.debug(f"Non lancés data: {len(non_lances_dict)} programmes")

        # Add data rows
        total_consistance = 0
        total_acheves = 0
        total_en_cours = 0
        total_non_lances = 0

        for i, (programme, consistance) in enumerate(programmes_situation):
            row = self._current_row + i
            self._logger.debug(f"Processing programme '{programme}' at row {row}")

            # Column A: Programme name
            sheet[f"A{row}"] = programme
            sheet[f"A{row}"].font = Font(name="Arial", size=9)

            # Column B: Consistance
            sheet[f"B{row}"] = consistance
            sheet[f"B{row}"].font = Font(name="Arial", size=9)
            total_consistance += consistance

            # Column C: Achevés
            acheves = acheves_dict.get(programme, 0)
            sheet[f"C{row}"] = acheves if acheves > 0 else "-"
            sheet[f"C{row}"].font = Font(name="Arial", size=9)
            total_acheves += acheves

            # Column D: En cours
            en_cours = en_cours_dict.get(programme, 0)
            sheet[f"D{row}"] = en_cours if en_cours > 0 else "-"
            sheet[f"D{row}"].font = Font(name="Arial", size=9)
            total_en_cours += en_cours

            # Column E: Non lancés
            non_lances = non_lances_dict.get(programme, 0)
            sheet[f"E{row}"] = non_lances if non_lances > 0 else "-"
            sheet[f"E{row}"].font = Font(name="Arial", size=9)
            total_non_lances += non_lances

            # Add borders
            for col in ["A", "B", "C", "D", "E"]:
                cell = sheet[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        # Add Total row for second table
        self._current_row += len(programmes_situation)
        self._logger.debug(
            f"Adding Total row for second table at row {self._current_row}"
        )

        sheet[f"A{self._current_row}"] = "Total général"
        sheet[f"B{self._current_row}"] = total_consistance
        sheet[f"C{self._current_row}"] = total_acheves
        sheet[f"D{self._current_row}"] = total_en_cours
        sheet[f"E{self._current_row}"] = total_non_lances

        # Format Total row
        for col in ["A", "B", "C", "D", "E"]:
            cell = sheet[f"{col}{self._current_row}"]
            cell.font = Font(name="Arial", size=9, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        self._logger.info(
            f"Second table completed with {len(programmes_situation)} programmes plus totals"
        )
        self._current_row += 2

    def _add_second_table_header(self, sheet: Worksheet) -> None:
        sheet[f"A{self._current_row}"] = (
            "Situation des programmes (à renseigner par la BNH, ex-CNL)"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        self._current_row += 1
        sheet[f"A{self._current_row}"] = (
            f"Arrêté le {self._document_context.report_date.strftime("%d/%m/%Y")}"
        )
        sheet.merge_cells(f"A{self._current_row}:E{self._current_row}")
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        self._current_row += 2

    def _add_footer(self, sheet: Worksheet) -> None:
        self._logger.debug("Adding document footer")

        # Left footer text (A-B)
        sheet.merge_cells(f"A{self._current_row}:B{self._current_row}")
        sheet[f"A{self._current_row}"] = "Visa du directeur régional de la BNH (ex-CNL)"
        sheet[f"A{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"A{self._current_row}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        # Right footer text (D-E)
        sheet.merge_cells(f"D{self._current_row}:E{self._current_row}")
        sheet[f"D{self._current_row}"] = "Visa du directeur du logement"
        sheet[f"D{self._current_row}"].font = Font(name="Arial", size=9, bold=True)
        sheet[f"D{self._current_row}"].alignment = Alignment(
            horizontal="right", vertical="center"
        )

        self._logger.debug("Footer added successfully")

    def _finalize_formatting(self, sheet: Worksheet) -> None:
        self._logger.debug("Applying final formatting")

        column_widths: dict[str, int] = {"A": 25, "B": 18, "C": 22, "D": 18, "E": 22}
        self._logger.debug(f"Setting column widths: {column_widths}")

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        self._logger.debug("Set page orientation to portrait with fit to width")
        self._logger.info("Final formatting completed successfully")
